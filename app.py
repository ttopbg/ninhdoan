import streamlit as st
import pandas as pd
import os
from io import BytesIO

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Thống kê thuế cơ sở",
    page_icon="🌚",
    layout="centered",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Be Vietnam Pro', sans-serif;
}

/* ── Background: clean light grey ── */
.stApp {
    background: #f2f5f9;
    min-height: 100vh;
}

/* ── Top header bar ── */
header[data-testid="stHeader"] {
    background: #ffffff;
    border-bottom: 3px solid #0077b6;
}

/* ── Title ── */
h1 {
    color: #05151f !important;
    font-weight: 800 !important;
    letter-spacing: -0.8px;
    text-align: center;
}

/* ── Subtitle ── */
.subtitle {
    text-align: center;
    color: #1e4060;
    font-size: 1rem;
    font-weight: 500;
    margin-bottom: 1.8rem;
    margin-top: -0.3rem;
}

/* ── White card ── */
.card {
    background: #ffffff;
    border: 1px solid #b8cfe0;
    border-radius: 14px;
    padding: 1.8rem 2.2rem;
    box-shadow: 0 2px 12px rgba(0,80,140,0.07);
    margin-bottom: 1.4rem;
}

/* ── Section headers h3 ── */
h3 {
    color: #05151f !important;
    font-weight: 700 !important;
    font-size: 1.05rem;
    margin-bottom: 0.6rem;
}

/* ── Body text / markdown ── */
p, li, .stMarkdown p {
    color: #0f2a3f !important;
    font-weight: 500;
    line-height: 1.75;
}

strong {
    color: #05151f !important;
    font-weight: 700;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #0096c7 !important;
    border-radius: 10px !important;
    background: #eef7fc !important;
    padding: 0.6rem !important;
}
[data-testid="stFileUploader"] p,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] div {
    color: #0a2e47 !important;
    font-weight: 600 !important;
}

/* ── Primary button ── */
.stButton > button {
    background: linear-gradient(90deg, #0077b6, #00b4d8) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    padding: 0.7rem 2rem !important;
    width: 100%;
    box-shadow: 0 4px 14px rgba(0,119,182,0.28);
    transition: box-shadow 0.2s, transform 0.15s;
}
.stButton > button:hover {
    box-shadow: 0 6px 20px rgba(0,119,182,0.4) !important;
    transform: translateY(-1px) !important;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(90deg, #00916e, #00c896) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    padding: 0.7rem 2rem !important;
    width: 100%;
    box-shadow: 0 4px 14px rgba(0,145,110,0.28);
}

/* ── Metric cards ── */
[data-testid="metric-container"] {
    background: #ffffff !important;
    border: 1px solid #a8c8e0 !important;
    border-radius: 12px !important;
    padding: 1rem !important;
    box-shadow: 0 2px 8px rgba(0,80,140,0.06) !important;
}
[data-testid="stMetricValue"] {
    color: #0055a5 !important;
    font-weight: 800 !important;
    font-size: 1.55rem !important;
}
[data-testid="stMetricLabel"] {
    color: #0f3050 !important;
    font-weight: 700 !important;
    font-size: 0.85rem !important;
}

/* ── Info / Success / Error alerts ── */
div[data-testid="stAlert"] p,
div[data-testid="stNotification"] p {
    font-weight: 600 !important;
    color: #0a2035 !important;
}

/* ── Spinner text ── */
[data-testid="stSpinner"] p {
    color: #0f2a3f !important;
    font-weight: 600 !important;
}

/* ── Divider ── */
hr { border-color: #b0c8dc !important; }

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border: 1px solid #b8cfe0 !important;
    border-radius: 10px !important;
    overflow: hidden;
}

/* ── Footer ── */
.footer-text {
    text-align: center;
    color: #3a607a;
    font-size: 0.8rem;
    font-weight: 500;
}
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.title("🎲 Thống kê thuế cơ sở theo Tỉnh")
st.markdown(
    '<p class="subtitle">Tải lên file Excel → nhận ngay báo cáo tổng hợp số cơ quan của mỗi tỉnh</p>',
    unsafe_allow_html=True,
)

# ── Upload card ───────────────────────────────────────────────────────────────
# st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("### 📁 Chọn file Excel đầu vào")
uploaded_file = st.file_uploader(
    label="Kéo thả hoặc Browse file (.xlsx / .xls)",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

col_info1, col_info2 = st.columns(2)
with col_info1:
    st.markdown("""
**Yêu cầu file:**
- Cột A: Tỉnh
- Cột B: Cơ quan
""")
with col_info2:
    st.markdown("""
**Xử lý tự động:**
- Bỏ qua dòng Tỉnh hoặc cơ sở trống
- Tổng hợp tại cột D & E
""")
# st.markdown('</div>', unsafe_allow_html=True)

# ── Process ───────────────────────────────────────────────────────────────────
if uploaded_file is not None:
    original_name = os.path.splitext(uploaded_file.name)[0]
    output_filename = f"{original_name}_output.xlsx"

    st.markdown('<div class="card">', unsafe_allow_html=True)
    start_btn = st.button("▶  Bắt đầu xử lý", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if start_btn:
        with st.spinner("⏳ Đang đọc và xử lý dữ liệu…"):
            try:
                df = pd.read_excel(uploaded_file, header=0, dtype=str)

                # Lấy 2 cột đầu tiên
                df = df.iloc[:, :2]
                df.columns = ["Tinh", "Co_quan"]
                total_rows = len(df)

                # Lọc dòng hợp lệ
                df_clean = df.dropna(subset=["Tinh", "Co_quan"])
                df_clean = df_clean[
                    (df_clean["Tinh"].str.strip() != "") &
                    (df_clean["Co_quan"].str.strip() != "")
                ].copy()
                skipped = total_rows - len(df_clean)

                # Tổng hợp số cơ quan duy nhất theo tỉnh
                summary = (
                    df_clean.groupby("Tinh", sort=True)["Co_quan"]
                    .nunique()
                    .reset_index()
                )
                summary.columns = ["Tỉnh", "Số cơ quan"]
                summary = summary.sort_values("Tỉnh").reset_index(drop=True)

                # Tạo file Excel output
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_clean.rename(columns={"Tinh": "Tỉnh", "Co_quan": "Cơ quan"}).to_excel(
                        writer, sheet_name="Sheet1", index=False, startrow=0, startcol=0
                    )
                    ws = writer.sheets["Sheet1"]

                    ws.cell(row=1, column=4, value="Tỉnh")
                    ws.cell(row=1, column=5, value="Số cơ quan")
                    for i, row in summary.iterrows():
                        ws.cell(row=i + 2, column=4, value=row["Tỉnh"])
                        ws.cell(row=i + 2, column=5, value=row["Số cơ quan"])

                    from openpyxl.utils import get_column_letter
                    for col_idx in range(1, 6):
                        max_len = 0
                        col_letter = get_column_letter(col_idx)
                        for cell in ws[col_letter]:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                output.seek(0)

                # ── Metrics ──────────────────────────────────────────────────
                st.markdown("---")
                st.markdown("### ✅ Kết quả xử lý")
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Tổng dòng", f"{total_rows:,}")
                m2.metric("Dòng hợp lệ", f"{len(df_clean):,}")
                m3.metric("Dòng bỏ qua", f"{skipped:,}")
                m4.metric("Số tỉnh", f"{len(summary):,}")

                # ── Preview ───────────────────────────────────────────────────
                st.markdown("### 📋 Bảng tổng hợp")
                st.dataframe(
                    summary,
                    use_container_width=True,
                    height=min(420, (len(summary) + 1) * 35 + 3),
                )

                # ── Download ──────────────────────────────────────────────────
                st.markdown("### 💾 Tải file kết quả")
                st.download_button(
                    label=f"⬇  Tải xuống  {output_filename}",
                    data=output,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success(f"Đã xử lý thành công! File output: **{output_filename}**")

            except Exception as e:
                st.error(f"❌ Lỗi khi xử lý file: {e}")
else:
    st.info("👆 Vui lòng tải lên file Excel để bắt đầu.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<p class="footer-text">Thống kê Cơ quan theo Tỉnh · Streamlit App</p>',
    unsafe_allow_html=True,
)
