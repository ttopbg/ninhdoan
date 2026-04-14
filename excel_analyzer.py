import streamlit as st
import pandas as pd
import os
from io import BytesIO

# ─── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Thống kê của Ninh Đoàng",
    page_icon="🍹",
    layout="centered",
)

# ─── CUSTOM CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@300;400;500;600;700&family=Playfair+Display:wght@600&display=swap');

/* ══════════════════════════════════════════
   CSS VARIABLES — LIGHT MODE (default)
   ══════════════════════════════════════════ */
:root {
    --bg-app:        #f0f4f8;
    --bg-app-mid:    #e8f0fe;
    --bg-app-end:    #f5f0ff;

    --bg-card:       #ffffff;
    --border-card:   rgba(200, 210, 240, 0.8);
    --shadow-card:   0 8px 32px rgba(60,80,160,0.10), 0 1.5px 4px rgba(60,80,160,0.05);

    --text-heading:  #1a2540;
    --text-sub:      #3d4f7a;
    --text-muted:    #5a6a90;
    --text-label:    #4a5a8a;

    --accent-blue:   #3a5fd4;
    --accent-purple: #6b46cc;

    --metric-bg:     linear-gradient(135deg, #dde8ff 0%, #ede5ff 100%);
    --metric-border: #c0cefc;
    --metric-val:    #2040a8;
    --metric-lbl:    #5060a0;

    --msg-info-bg:   #ddeaff;
    --msg-info-bdr:  #4a6ee0;
    --msg-info-txt:  #1a3490;

    --msg-ok-bg:     #d4f5e8;
    --msg-ok-bdr:    #1fa870;
    --msg-ok-txt:    #0d5c3a;

    --msg-warn-bg:   #fff0d0;
    --msg-warn-bdr:  #d48a00;
    --msg-warn-txt:  #7a4a00;

    --upload-bg:     rgba(210,228,255,0.40);
    --upload-bdr:    #90aae8;
    --upload-bdr-hv: #4a6ee0;

    --divider:       #c8d4f0;
    --table-border:  #c0ceee;

    --section-lbl:   #3d5090;
}

/* ══════════════════════════════════════════
   CSS VARIABLES — DARK MODE
   ══════════════════════════════════════════ */
@media (prefers-color-scheme: dark) {
    :root {
        --bg-app:        #0d1117;
        --bg-app-mid:    #111827;
        --bg-app-end:    #130d24;

        --bg-card:       #1a2035;
        --border-card:   rgba(80, 100, 180, 0.35);
        --shadow-card:   0 8px 32px rgba(0,0,0,0.40), 0 1.5px 4px rgba(0,0,0,0.30);

        --text-heading:  #e8eeff;
        --text-sub:      #b8c8f0;
        --text-muted:    #8898c8;
        --text-label:    #a0b0e0;

        --accent-blue:   #6a9aff;
        --accent-purple: #a07aff;

        --metric-bg:     linear-gradient(135deg, #1e2d56 0%, #221845 100%);
        --metric-border: #2d3f78;
        --metric-val:    #7ab0ff;
        --metric-lbl:    #8898c0;

        --msg-info-bg:   #172040;
        --msg-info-bdr:  #5070d0;
        --msg-info-txt:  #90b8ff;

        --msg-ok-bg:     #0d2820;
        --msg-ok-bdr:    #28a870;
        --msg-ok-txt:    #60e0a8;

        --msg-warn-bg:   #281800;
        --msg-warn-bdr:  #c07800;
        --msg-warn-txt:  #ffc040;

        --upload-bg:     rgba(30, 50, 100, 0.40);
        --upload-bdr:    #3a5090;
        --upload-bdr-hv: #6a9aff;

        --divider:       #2a3560;
        --table-border:  #2a3560;

        --section-lbl:   #8898d0;
    }
}

/* ══════════════════════════════════════════
   Streamlit dark-theme class override
   (khi người dùng chọn dark trong Streamlit)
   ══════════════════════════════════════════ */
[data-theme="dark"] {
    --bg-app:        #0d1117;
    --bg-app-mid:    #111827;
    --bg-app-end:    #130d24;
    --bg-card:       #1a2035;
    --border-card:   rgba(80, 100, 180, 0.35);
    --shadow-card:   0 8px 32px rgba(0,0,0,0.40), 0 1.5px 4px rgba(0,0,0,0.30);
    --text-heading:  #e8eeff;
    --text-sub:      #b8c8f0;
    --text-muted:    #8898c8;
    --text-label:    #a0b0e0;
    --accent-blue:   #6a9aff;
    --accent-purple: #a07aff;
    --metric-bg:     linear-gradient(135deg, #1e2d56 0%, #221845 100%);
    --metric-border: #2d3f78;
    --metric-val:    #7ab0ff;
    --metric-lbl:    #8898c0;
    --msg-info-bg:   #172040;
    --msg-info-bdr:  #5070d0;
    --msg-info-txt:  #90b8ff;
    --msg-ok-bg:     #0d2820;
    --msg-ok-bdr:    #28a870;
    --msg-ok-txt:    #60e0a8;
    --msg-warn-bg:   #281800;
    --msg-warn-bdr:  #c07800;
    --msg-warn-txt:  #ffc040;
    --upload-bg:     rgba(30, 50, 100, 0.40);
    --upload-bdr:    #3a5090;
    --upload-bdr-hv: #6a9aff;
    --divider:       #2a3560;
    --table-border:  #2a3560;
    --section-lbl:   #8898d0;
}

/* ─── Base ─── */
html, body, [class*="css"] {
    font-family: 'Be Vietnam Pro', sans-serif;
}

.stApp {
    background: linear-gradient(145deg,
        var(--bg-app) 0%,
        var(--bg-app-mid) 50%,
        var(--bg-app-end) 100%) !important;
    min-height: 100vh;
}

#MainMenu, footer, header { visibility: hidden; }

/* ─── Hero ─── */
.hero {
    text-align: center;
    padding: 2.5rem 1rem 1.5rem;
}
.hero h1 {
    font-family: 'Playfair Display', serif;
    font-size: 2.4rem;
    color: var(--text-heading) !important;
    margin-bottom: 0.3rem;
    letter-spacing: -0.5px;
}
.hero h4 {
    color: var(--text-sub) !important;
    font-size: 1rem;
    font-weight: 400;
    margin: 0;
}

/* ─── Card ─── */
.card {
    background: var(--bg-card) !important;
    border: 1px solid var(--border-card);
    border-radius: 20px;
    padding: 2rem 2rem 1.6rem;
    box-shadow: var(--shadow-card);
    margin-bottom: 1.4rem;
}

/* ─── Section label ─── */
.section-label {
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 1.8px;
    text-transform: uppercase;
    color: var(--section-lbl) !important;
    margin-bottom: 0.6rem;
}

/* ─── File uploader ─── */
[data-testid="stFileUploader"] {
    border: 2px dashed var(--upload-bdr) !important;
    border-radius: 14px !important;
    background: var(--upload-bg) !important;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: var(--upload-bdr-hv) !important;
}
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] p,
[data-testid="stFileUploader"] span {
    color: var(--text-label) !important;
    font-weight: 500 !important;
}

/* ─── Buttons ─── */
.stButton > button {
    background: linear-gradient(135deg, var(--accent-blue) 0%, var(--accent-purple) 100%) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.75rem 2.2rem !important;
    font-family: 'Be Vietnam Pro', sans-serif !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.3px !important;
    box-shadow: 0 4px 18px rgba(79,114,227,0.35) !important;
    transition: all 0.2s ease !important;
    width: 100% !important;
}
.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(79,114,227,0.50) !important;
    filter: brightness(1.1) !important;
}
.stButton > button:active { transform: translateY(0px) !important; }
.stButton > button:disabled {
    opacity: 0.45 !important;
    filter: grayscale(30%) !important;
}

[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #22c07a 0%, #159a5c 100%) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.75rem 2rem !important;
    font-family: 'Be Vietnam Pro', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    box-shadow: 0 4px 18px rgba(34,192,122,0.35) !important;
    width: 100% !important;
}
[data-testid="stDownloadButton"] > button:hover {
    transform: translateY(-2px) !important;
    filter: brightness(1.12) !important;
    box-shadow: 0 8px 24px rgba(34,192,122,0.50) !important;
}

/* ─── Metric boxes ─── */
.metrics-row {
    display: flex;
    gap: 1rem;
    margin: 1.2rem 0 0.6rem;
    flex-wrap: wrap;
}
.metric-box {
    flex: 1;
    min-width: 120px;
    background: var(--metric-bg);
    border: 1px solid var(--metric-border);
    border-radius: 14px;
    padding: 1rem 1.2rem;
    text-align: center;
}
.metric-box .val {
    font-size: 1.9rem;
    font-weight: 700;
    color: var(--metric-val) !important;
    line-height: 1.1;
}
.metric-box .lbl {
    font-size: 0.78rem;
    color: var(--metric-lbl) !important;
    font-weight: 500;
    margin-top: 0.2rem;
    letter-spacing: 0.4px;
}

/* ─── Status messages ─── */
.msg-info {
    background: var(--msg-info-bg) !important;
    border-left: 4px solid var(--msg-info-bdr);
    border-radius: 8px;
    padding: 0.8rem 1rem;
    color: var(--msg-info-txt) !important;
    font-size: 0.9rem;
    margin: 0.8rem 0;
}
.msg-info strong { color: var(--msg-info-txt) !important; }

.msg-success {
    background: var(--msg-ok-bg) !important;
    border-left: 4px solid var(--msg-ok-bdr);
    border-radius: 8px;
    padding: 0.8rem 1rem;
    color: var(--msg-ok-txt) !important;
    font-size: 0.9rem;
    margin: 0.8rem 0;
}
.msg-success strong { color: var(--msg-ok-txt) !important; }

.msg-warn {
    background: var(--msg-warn-bg) !important;
    border-left: 4px solid var(--msg-warn-bdr);
    border-radius: 8px;
    padding: 0.8rem 1rem;
    color: var(--msg-warn-txt) !important;
    font-size: 0.9rem;
    margin: 0.8rem 0;
}
.msg-warn strong { color: var(--msg-warn-txt) !important; }

/* ─── Table ─── */
[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    border: 1px solid var(--table-border) !important;
}

/* ─── Divider ─── */
hr {
    border: none;
    border-top: 1px solid var(--divider);
    margin: 1.2rem 0;
}

/* ─── Streamlit default text colour fix ─── */
.stMarkdown p, .stMarkdown span, .stMarkdown div {
    color: var(--text-sub);
}
</style>
""", unsafe_allow_html=True)


# ─── HERO ────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
    <h1>🧋 Thống kê theo Tỉnh</h1>
    <h4>Tổng hợp dữ liệu thuế cơ sở & Doanh nghiệp</h4>
</div>
""", unsafe_allow_html=True)


# ─── UPLOAD CARD ─────────────────────────────────────────────────────────────────
# st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-label">📁 Tải lên file Excel</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    label="Chọn file Excel (.xlsx hoặc .xls)",
    type=["xlsx", "xls"],
    help="File cần có 3 cột: Tỉnh (A), Cơ sở (B), Doanh nghiệp (C)",
    label_visibility="collapsed",
)

if uploaded_file:
    st.markdown(f"""
    <div class="msg-info">
        ✅ Đã tải lên: <strong>{uploaded_file.name}</strong>
        &nbsp;|&nbsp; Kích thước: <strong>{uploaded_file.size / 1024:.1f} KB</strong>
    </div>
    """, unsafe_allow_html=True)

# st.markdown('</div>', unsafe_allow_html=True)


# ─── PROCESS CARD ────────────────────────────────────────────────────────────────
# st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-label">⚙️ Xử lý dữ liệu</div>', unsafe_allow_html=True)

run_btn = st.button("🚀 Bắt đầu thống kê", disabled=(uploaded_file is None))

if run_btn and uploaded_file:
    with st.spinner("Đang xử lý dữ liệu, vui lòng chờ..."):
        try:
            # ── Read file ──
            file_bytes = uploaded_file.read()
            df_raw = pd.read_excel(BytesIO(file_bytes), header=0, dtype=str)

            # ── Normalise columns A, B, C ──
            df_raw.columns = [str(c).strip() for c in df_raw.columns]

            # Grab first 3 columns regardless of header name
            col_tinh = df_raw.columns[0]
            col_co_so = df_raw.columns[1]
            col_dn = df_raw.columns[2]

            total_rows = len(df_raw)

            # ── Drop rows where any of the 3 key columns is blank / NaN ──
            df = df_raw[[col_tinh, col_co_so, col_dn]].copy()
            df = df.replace(r'^\s*$', pd.NA, regex=True)
            df = df.dropna(subset=[col_tinh, col_co_so, col_dn])
            dropped = total_rows - len(df)

            # ── Build summary: đếm số dòng theo tỉnh ──
            # Số cơ sở  = số dòng có giá trị cột B thuộc tỉnh đó
            # Số doanh nghiệp = số dòng có giá trị cột C thuộc tỉnh đó
            summary = (
                df.groupby(col_tinh, sort=True)
                  .agg(
                      so_co_so=(col_co_so, 'count'),
                      so_dn=(col_dn, 'count'),
                  )
                  .reset_index()
            )
            summary.columns = ['Tỉnh', 'Số Cơ Sở', 'Số Doanh Nghiệp']

            # ── Build output: chỉ đúng 3 cột gốc + D trống + E/F/G thống kê ──
            df_out = df_raw[[col_tinh, col_co_so, col_dn]].copy()
            df_out.columns = ['Tỉnh', 'Cơ Sở', 'Doanh Nghiệp']

            n_rows    = len(df_out)
            n_summary = len(summary)

            # Chuẩn bị giá trị cột E, F, G (chỉ điền đến hết summary, còn lại None)
            e_vals = [None] * n_rows
            f_vals = [None] * n_rows
            g_vals = [None] * n_rows
            for i in range(min(n_summary, n_rows)):
                e_vals[i] = summary.iloc[i]['Tỉnh']
                f_vals[i] = int(summary.iloc[i]['Số Cơ Sở'])
                g_vals[i] = int(summary.iloc[i]['Số Doanh Nghiệp'])

            # ── Save to BytesIO — ghi thủ công bằng openpyxl để kiểm soát cột D ──
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'Data'

            # ---- Header row ----
            headers = {
                1: ('Tỉnh',          '4F72E3'),
                2: ('Cơ Sở',         '4F72E3'),
                3: ('Doanh Nghiệp',  '4F72E3'),
                4: (None,            None),       # D trống
                5: ('Tỉnh (TH)',     '7B5CDE'),
                6: ('Số Cơ Sở',      '7B5CDE'),
                7: ('Số DN',         '7B5CDE'),
            }
            hdr_font  = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
            ctr_align = Alignment(horizontal='center', vertical='center')

            for col_idx, (label, color) in headers.items():
                cell = ws.cell(row=1, column=col_idx, value=label)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = hdr_font
                    cell.alignment = ctr_align

            # ---- Data rows ----
            alt_fill     = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
            sum_data_fill= PatternFill(start_color='F5F0FF', end_color='F5F0FF', fill_type='solid')
            thin_border  = Border(
                left=Side(style='thin', color='DDEAFF'),
                right=Side(style='thin', color='DDEAFF'),
                bottom=Side(style='thin', color='DDEAFF'),
            )

            # Ghi nhanh toàn bộ dữ liệu bằng ws.append (tối ưu cho 60k dòng)
            tinh_vals_list = df_out['Tỉnh'].tolist()
            coso_vals_list = df_out['Cơ Sở'].tolist()
            dn_vals_list   = df_out['Doanh Nghiệp'].tolist()

            for i in range(n_rows):
                ws.append([tinh_vals_list[i], coso_vals_list[i], dn_vals_list[i],
                           None, e_vals[i], f_vals[i], g_vals[i]])

            # Áp style cho ~500 dòng đầu (style toàn bộ 60k dòng sẽ rất chậm)
            style_limit = min(n_rows + 2, 502)
            for r in range(2, style_limit):
                for col_idx in range(1, 8):
                    if col_idx == 4:
                        continue
                    cell = ws.cell(row=r, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    if r % 2 == 0:
                        cell.fill = sum_data_fill if col_idx >= 5 else alt_fill

            # ---- Column widths ----
            for col_idx in range(1, 8):
                col_letter = get_column_letter(col_idx)
                if col_idx == 4:
                    ws.column_dimensions[col_letter].width = 3
                    continue
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or ''))
                     for r in range(1, min(n_rows + 2, 500))),
                    default=10
                )
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

            ws.freeze_panes = 'A2'

            output_buffer = BytesIO()
            wb.save(output_buffer)

            output_buffer.seek(0)
            # ── Output filename ──
            base_name = os.path.splitext(uploaded_file.name)[0]
            output_filename = f"{base_name}_output.xlsx"

            # ── Store in session state ──
            st.session_state['output_data']     = output_buffer.getvalue()
            st.session_state['output_filename'] = output_filename
            st.session_state['summary']         = summary
            st.session_state['total_rows']      = total_rows
            st.session_state['dropped']         = dropped
            st.session_state['valid_rows']      = len(df)
            st.session_state['num_provinces']   = len(summary)

        except Exception as e:
            st.markdown(f'<div class="msg-warn">❌ Lỗi xử lý file: <strong>{e}</strong></div>', unsafe_allow_html=True)

# st.markdown('</div>', unsafe_allow_html=True)


# ─── RESULTS ─────────────────────────────────────────────────────────────────────
if 'output_data' in st.session_state:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-label">📈 Kết quả thống kê</div>', unsafe_allow_html=True)

    # Metrics
    st.markdown(f"""
    <div class="metrics-row">
        <div class="metric-box">
            <div class="val">{st.session_state['total_rows']:,}</div>
            <div class="lbl">Tổng dòng gốc</div>
        </div>
        <div class="metric-box">
            <div class="val">{st.session_state['dropped']:,}</div>
            <div class="lbl">Dòng bỏ qua (trống)</div>
        </div>
        <div class="metric-box">
            <div class="val">{st.session_state['valid_rows']:,}</div>
            <div class="lbl">Dòng hợp lệ</div>
        </div>
        <div class="metric-box">
            <div class="val">{st.session_state['num_provinces']}</div>
            <div class="lbl">Số tỉnh duy nhất</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<hr>', unsafe_allow_html=True)

    # Summary table
    st.markdown('<div class="section-label">🗂 Bảng tổng hợp theo Tỉnh</div>', unsafe_allow_html=True)
    st.dataframe(
        st.session_state['summary'].style.format({'Số Cơ Sở': '{:,}', 'Số Doanh Nghiệp': '{:,}'}),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown('<hr>', unsafe_allow_html=True)

    # Download
    st.markdown('<div class="section-label">💾 Tải file kết quả</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="msg-success">
        ✅ File đã sẵn sàng: <strong>{st.session_state['output_filename']}</strong>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label="⬇️  Tải xuống file Excel",
        data=st.session_state['output_data'],
        file_name=st.session_state['output_filename'],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown('</div>', unsafe_allow_html=True)


# # ─── FOOTER ──────────────────────────────────────────────────────────────────────
# st.markdown("""
# <div style="text-align:center; color:#b0b8d0; font-size:0.78rem; padding: 1.5rem 0 0.5rem; font-weight:300;">
#     Thống Kê Tỉnh · Được xây dựng với Streamlit &amp; Pandas
# </div>
# """, unsafe_allow_html=True)
